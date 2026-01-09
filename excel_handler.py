import os
import xlwings as xw
from openpyxl import load_workbook
from datetime import datetime, timedelta
from config import EXCEL_PATH


def get_keyword_from_xlsm():
    if not os.path.exists(EXCEL_PATH):
        print(f"파일을 찾을 수 없습니다: {EXCEL_PATH}")
        return set()

    app = xw.App(visible=False)  # 키워드만 뽑을 때는 백그라운드 실행
    app.display_alerts = False  # 팝업 알림 차단

    try:
        # 1. 엑셀 로드
        # # # keep_vba=True: 매크로 유지
        # # # data_only=True: 수식이 아닌 '텍스트 결과값'만 가져옴
        # # wb = load_workbook(EXCEL_PATH, keep_vba=True, data_only=True)
        # # ws = wb['데이터']
        # wb = xw.Book(EXCEL_PATH)

        # [수정] update_links=False로 팝업을 띄우지 않고 열기
        wb = app.books.open(EXCEL_PATH, update_links=False)

        # [수정] 열린 직후 코드로 직접 외부 링크 업데이트 명령
        try:
            # 1은 엑셀 통합 문서 링크를 의미합니다.
            links = wb.api.LinkSources(1)
            if links is not None:
                # 리스트가 비어있지 않은지 확인 후 업데이트
                wb.api.UpdateLink(Name=links)
                print("외부 연결 데이터 업데이트 완료.")
        except Exception as update_err:
            # 업데이트 중 오류가 나도 키워드 추출은 계속 진행하도록 처리
            print(f"외부 링크 업데이트 건너뜀 (사유: {update_err})")

        ws = wb.sheets['데이터']

        # 2. 키워드 가져오기
        # keywords = set()
        # for row in ws.iter_rows(min_row=7, min_col=10, max_col=10):
        #     cell_value = row[0].value # iter_rows는 한 행을 셀들의 묶음(튜플)으로 반환
        #
        #     if cell_value is None:
        #         continue
        #
        #     keyword = str(cell_value).strip()
        #
        #     if not keyword:
        #         continue
        #
        #     keywords.add(keyword)

        # J열(10번째) 7행부터 마지막 데이터가 있는 행까지 한 번에 가져오기
        last_row = ws.range('J' + str(ws.api.Rows.Count)).end('up').row
        if last_row < 7:
            wb.close()
            return set()

        # 7행부터 시트 전체 마지막 행까지의 J열(10번째) 데이터를 가져옴
        values = ws.range((7, 10), (last_row, 10)).value

        if not isinstance(values, list):
            values = [values]

        keywords = set()
        for cell_value in values:
            if cell_value:
                keywords.add(str(cell_value).strip())

        print(f"키워드 추출: {list(keywords)}")

        wb.close()
        return keywords

    finally:
            # 작업 완료 후 앱 종료
            app.quit()



# 2026-01-01부터 날짜 열 확인 및 추가
def sync_date_columns_until_today(ws, start_date_str="2026-01-01"):
    """
    1월 1일부터 오늘까지 누락된 날짜 열을 5행에 자동으로 추가
    고정 필드를 유지하기 위해 열을 삽입(Insert)하며 확장
    """
    # 날짜 범위 설정
    start_date = datetime.strptime(start_date_str, "%Y-%m-%d")
    today = datetime.now()

    # 날짜를 넣기 시작할 시작 열 (BV = 74)
    start_search_col = 74

    # 1월 1일부터 오늘까지 반복하며 날짜 확인
    current_date = start_date
    while current_date <= today:
        date_header = f"{current_date.month}/{current_date.day}"
        is_exist = False    # 5행에서 해당 날짜가 이미 있는지 확인

        # # cell_val = ws.cell(row=5, column=74).value
        #
        # # 현재 5행에 해당 날짜가 이미 있는지 끝까지 검사
        # last_col = ws.max_column
        # for col in range(start_search_col, last_col+1):
        #     cell_val = ws.cell(row=5, column=col).value
        #     if cell_val is None:
        #         continue
        #
        #     # 날짜 (객체 또는 문자열) 비교
        #     if isinstance(cell_val, datetime):
        #         check_str = f"{cell_val.month}/{cell_val.day}"
        #     else:
        #         check_str = str(cell_val).strip()
        #
        #     if check_str == date_header:
        #         is_exist = True
        #         break

        # 5행의 헤더들을 리스트로 가져와서 비교 (속도 최적화)
        last_col = ws.range(5, ws.api.Columns.Count).end('left').column
        if last_col < start_search_col: last_col = start_search_col

        row_5_values = ws.range((5, start_search_col), (5, last_col + 5)).value

        if row_5_values:
            for val in row_5_values:
                if val is None: continue
                check_str = f"{val.month}/{val.day}" if isinstance(val, datetime) else str(val).strip()
                if check_str == date_header:
                    is_exist = True
                    break

        # 날짜가 없으면 "가장 오른쪽 끝" 혹은 "특정 고정필드 직전"에 추가
        if not is_exist:
            # '비고'나 '서식' 같은 고정 헤더가 시작되는 위치를 찾음
            target_col = start_search_col
            while True:
                # val = ws.cell(row=5, column=target_col).value
                val = ws.range(5, target_col).value
                # 빈칸이거나 고정 키워드가 나오면 그 자리에 삽입
                if val is None or any(kw in str(val) for kw in ["직전", "비고", "서식", "공란"]):
                    break
                target_col += 1

            # ws.insert_cols(target_col)
            # ws.cell(row=5, column=target_col).value = date_header

            # xlwings의 api를 사용하여 엑셀 열 삽입 기능 호출
            ws.range((1, target_col)).api.EntireColumn.Insert()
            ws.range(5, target_col).value = date_header
            print(f"새 열 추가됨: {target_col}번째 열 ({date_header})")

        current_date += timedelta(days=1)



def get_all_date_texts_from_header(ws):
    """
    엑셀 5행의 BV열(74)부터 고정 필드 전까지의 날짜들을
    ['2026-01-01', '2026-01-02', ...] 형태의 텍스트 리스트로 반환
    """
    COL_BV = 74
    date_text_list = []

    # [수정] 5행 기준 마지막 열 찾기 (ws.max_column 대체)
    max_col = ws.range(5, ws.api.Columns.Count).end('left').column

    # BV열(74)부터 오른쪽으로 하나씩 검사
    for col in range(COL_BV, max_col + 1):
        # cell_val = ws.cell(row=5, column=col).value
        cell_val = ws.range(5, col).value

        # 빈 칸이거나 날짜가 아닌 고정 헤더를 만나면 탐색 중단
        if cell_val is None:
            break

        cell_str = str(cell_val).strip()
        if any(kw in cell_str for kw in ["직전", "비고", "서식", "공란"]):
            break

        # 날짜 데이터를 텍스트로 변환
        formatted_date = None

        # 셀 값이 datetime 객체인 경우
        if isinstance(cell_val, datetime):
            formatted_date = cell_val.strftime('%Y-%m-%d')

        # 셀 값이 '1/1' 또는 '2026-01-01' 같은 문자열인 경우
        else:
            try:
                # 이미 'YYYY-MM-DD' 형식인지 확인
                datetime.strptime(cell_str, '%Y-%m-%d')
                formatted_date = cell_str
            except ValueError:
                # '1/1' 형식인 경우 현재 연도를 붙여서 변환
                try:
                    current_year = datetime.now().year
                    dt = datetime.strptime(f"{current_year}/{cell_str}", "%Y/%m/%d")
                    formatted_date = dt.strftime('%Y-%m-%d')
                except:
                    # 형식을 알 수 없으면 그대로 문자열로 저장
                    formatted_date = cell_str

        if formatted_date:
            date_text_list.append(formatted_date)

    return date_text_list


def get_missing_dates_for_keyword(ws, keyword, date_info_map):
    """지정된 키워드에 대해 순위값이 비어있는 날짜들만 반환"""
    missing_dates = set()

    # [수정] J열 기준 마지막 행 찾기 (ws.max_row 대체)
    last_row = ws.range('J' + str(ws.api.Rows.Count)).end('up').row

    for row in range(7, last_row + 1):
        ex_kw = str(ws.range(row, 10).value or "").strip()
        if ex_kw == keyword:
            for d_str, col_idx in date_info_map.items():
                if ws.range(row, col_idx).value is None:
                    missing_dates.add(d_str)
    return list(missing_dates)


def update_excel_rank(ws, target_vi_id, target_keyword, rank_value, target_date):
    # 5행에서 날짜에 해당하는 열 번호 찾기
    target_col = None

    # 날짜 입력 형식을 안전하게 변환
    dt = datetime.strptime(target_date, '%Y-%m-%d')
    search_header = f"{dt.month}/{dt.day}"

    # for col in range(74, ws.max_column + 1):
    #     cell_val = ws.cell(row=5, column=col).value
    #     if cell_val is None: continue
    #
    #     # 엑셀 헤더가 날짜 객체인지 텍스트인지 판별하여 비교
    #     header = f"{cell_val.month}/{cell_val.day}" if isinstance(cell_val, datetime) else str(cell_val).strip()
    #
    #     if header == search_header:
    #         target_col = col
    #         break

    # [수정] 날짜 열 찾기 (리스트 내 탐색)
    row_5_values = ws.range((5, 74), (5, 200)).value
    target_col = None
    for i, val in enumerate(row_5_values):
        if val is None: continue
        header = f"{val.month}/{val.day}" if isinstance(val, datetime) else str(val).strip()
        if header == search_header:
            target_col = 74 + i
            break

    if not target_col:  # target_col이 여전히 None(False)일때
        print(f"엑셀에서 {search_header} 열을 찾지 못했습니다.")
        return

    # [수정] ID(6열)와 키워드(10열) 전체 데이터를 가져와서 행 매칭 (반복문 내 셀 접근 최소화)
    last_row = ws.range('J' + str(ws.cells.last_cell.row)).end('up').row
    ids = ws.range((7, 6), (last_row, 6)).value
    kws = ws.range((7, 10), (last_row, 10)).value

    if not isinstance(ids, list): ids = [ids]
    if not isinstance(kws, list): kws = [kws]

    for i, (raw_vi_id, kw) in enumerate(zip(ids, kws)):
        try:
            vi_id = str(int(float(raw_vi_id))) if raw_vi_id else ""
        except:
            vi_id = str(raw_vi_id or "").strip()

        if vi_id == str(target_vi_id) and str(kw).strip() == target_keyword:
            target_cell = ws.range(7 + i, target_col)

            # [중복 방지] 이미 동일한 값이 엑셀에 들어있다면 업데이트 생략
            if str(target_cell.value) == str(rank_value):
                return True  # 성공으로 간주하되 작업은 안 함

            target_cell.value = rank_value
            print(f"성공: {7 + i}행 {target_col}열에 '{rank_value}' 입력")
            return True

    return False  # 끝까지 못 찾은 경우

    # COL_VI_ID = 6  # F열
    # COL_KEYWORD = 10  # J열
    #
    # found = False
    # for row in range(7, ws.max_row + 1):
    #     # [수정] ID 비교 시 float 형태(.0)가 생기지 않도록 정수형 처리 후 문자열 변환
    #     raw_vi_id = ws.cell(row=row, column=COL_VI_ID).value
    #     try:
    #         # 12345.0 같은 데이터를 '12345'로 변환
    #         vi_id = str(int(float(raw_vi_id))) if raw_vi_id else ""
    #     except:
    #         vi_id = str(raw_vi_id or "").strip()
    #
    #     keyword = str(ws.cell(row=row, column=COL_KEYWORD).value or "").strip()
    #
    #     # ID와 키워드 동시 비교
    #     if vi_id == str(target_vi_id) and keyword == target_keyword:
    #         ws.cell(row=row, column=target_col).value = rank_value
    #         print(f"성공: {row}행 {target_col}열에 '{rank_value}' 입력")
    #         found = True
    #         break
    #
    # if not found:
    #     # 디버깅을 위해 찾지 못한 정보 출력
    #     pass


# 이 함수를 루프 밖에서 한 번만 호출하여 인덱스를 만듭니다.
def build_row_index(ws):
    last_row = ws.range('J' + str(ws.api.Rows.Count)).end('up').row
    # F열(ID)부터 J열(키워드)까지 한 번에 읽기
    data = ws.range((7, 6), (last_row, 10)).value

    row_index = {}
    for i, row in enumerate(data):
        raw_id, kw = row[0], row[4]  # F열은 0번, J열은 4번 인덱스
        try:
            vi_id = str(int(float(raw_id))) if raw_id else ""
        except:
            vi_id = str(raw_id or "").strip()

        # 키 생성 (ID, 키워드) -> 실제 행 번호는 i + 7
        row_index[(vi_id, str(kw).strip())] = i + 7
    return row_index


# 수정된 update 함수 (인덱스 사용)
def update_excel_rank_fast(ws, row_index, target_col, target_vi_id, target_keyword, rank_value):
    target_key = (str(target_vi_id), str(target_keyword).strip())
    row_num = row_index.get(target_key)

    if row_num:
        target_cell = ws.range(row_num, target_col)
        if str(target_cell.value) != str(rank_value):
            target_cell.value = rank_value
            return True
    return False



def get_dates_requiring_update(ws):
    """
    5행의 날짜 열들을 순회하며, 해당 열 전체(7행~마지막행)에
    순위 데이터가 '단 하나도' 없는 날짜 리스트만 반환합니다.
    """
    COL_BV = 74
    dates_to_update = []

    # 헤더 및 마지막 행 번호 파악
    max_col = ws.range(5, ws.api.Columns.Count).end('left').column
    if max_col < COL_BV: max_col = COL_BV

    headers = ws.range((5, COL_BV), (5, max_col + 1)).value
    # J열 기준으로 데이터가 있는 마지막 행 번호 파악
    last_row = ws.range('J' + str(ws.api.Rows.Count)).end('up').row

    # [수정] Q열(17번째 열)의 행 구분 텍스트를 가져옵니다.
    row_types = ws.range((7, 17), (last_row, 17)).value
    if not isinstance(row_types, list): row_types = [row_types]

    # for col in range(COL_BV, ws.max_column + 1):
    #     header_val = ws.cell(row=5, column=col).value
    for i, header_val in enumerate(headers):
        if header_val is None: break

        header_str = str(header_val).strip()
        # 고정 헤더를 만나면 탐색 중단
        if any(kw in header_str for kw in ["직전", "비고", "서식", "공란"]):
            break

        # has_any_data = False
        # for row in range(7, ws.max_row + 1):
        #     rank_val = ws.cell(row=row, column=col).value
        #
        #     # 셀 값이 None이 아니고, 공백 제외 문자열이 존재하면 데이터가 있는 것으로 간주
        #     if rank_val is not None and str(rank_val).strip() != "":
        #         has_any_data = True
        #         break  # 하나라도 데이터가 있으면 이 날짜는 검사 종료

        col_idx = COL_BV + i

        # [수정] 해당 열 전체 데이터를 리스트로 가져와서 빈칸 여부 검사
        col_data = ws.range((7, col_idx), (last_row, col_idx)).value
        if not isinstance(col_data, list): col_data = [col_data]

        # [핵심 수정 로직]
        # 해당 열에서 '순위'라고 적힌 행들에 대해서만 데이터가 있는지 확인
        # 판매수량, 광고, 가구매 행에 데이터가 있어도 '순위' 행들이 비어있으면 수집 대상에 포함
        has_actual_rank_data = False
        for idx, rank_val in enumerate(col_data):
            # Q열의 텍스트를 확인하여 '순위' 포함 여부 체크
            row_type = str(row_types[idx]).strip() if idx < len(row_types) and row_types[idx] else ""

            # 행 타입에 '순위'라는 글자가 포함된 행만 순위 데이터 존재 여부를 검사합니다.
            if "순위" in row_type:
                clean_val = str(rank_val).strip() if rank_val is not None else ""
                # 비어있지 않으면서 '0', '-', 'None'이 아닌 실제 순위 숫자가 있는지 확인
                if clean_val not in ["", "0", "0.0", "-", "None"]:
                    has_actual_rank_data = True
                    break

        # 순위 데이터가 '하나도 없는' 날짜만 업데이트 대상으로 선정
        if not has_actual_rank_data:
            if isinstance(header_val, datetime):
                formatted_date = header_val.strftime('%Y-%m-%d')
            else:
                try:
                    current_year = datetime.now().year
                    dt = datetime.strptime(f"{current_year}/{header_str}", "%Y/%m/%d")
                    formatted_date = dt.strftime('%Y-%m-%d')
                except:
                    formatted_date = header_str
            dates_to_update.append(formatted_date)

    # if dates_to_update:
    #     print(f"데이터가 완전히 비어 있는 날짜(크롤링 대상): {dates_to_update}")
    # else:
    #     print("모든 날짜 열에 최소 하나 이상의 데이터가 기록되어 있습니다.")

    return dates_to_update